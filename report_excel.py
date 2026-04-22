"""
report_excel.py — Gerador de relatório Excel (.xlsx) da conciliação
====================================================================
Gera um arquivo .xlsx com 5 abas:

  1. Resumo          — totais, percentuais e painel visual
  2. Conciliadas     — NFs que batem perfeitamente
  3. Divergentes     — NFs com algum campo diferente (valor A × valor B)
  4. Só Prefeitura   — NFs que existem na prefeitura mas não no sistema
  5. Só Sistema      — NFs que estão no sistema mas não na prefeitura

Dependências: openpyxl
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from comparator import ResultadoComparacao, STATUS_DIVERGENCIA_CRITICA, STATUS_DIVERGENCIA_LEVE


# ---------------------------------------------------------------------------
# Paleta de cores
# ---------------------------------------------------------------------------

COR_VERDE_ESCURO  = '1A6B3C'
COR_VERDE_CLARO   = 'D6F0E0'
COR_VERMELHO      = 'C0392B'
COR_VERMELHO_CLARO= 'FADBD8'
COR_AMARELO       = 'F39C12'
COR_AMARELO_CLARO = 'FEF9E7'
COR_AZUL          = '1A5276'
COR_AZUL_CLARO    = 'D6EAF8'
COR_CINZA_HEADER  = '2C3E50'
COR_CINZA_CLARO   = 'F2F3F4'
COR_BRANCO        = 'FFFFFF'
COR_LINHA_ALT     = 'F8F9FA'


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _fonte(bold=False, tamanho=10, cor='000000', nome='Arial'):
    return Font(name=nome, bold=bold, size=tamanho, color=cor)

def _fill(cor_hex):
    return PatternFill('solid', start_color=cor_hex, fgColor=cor_hex)

def _borda_fina():
    lado = Side(style='thin', color='CCCCCC')
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _alinhar(horizontal='left', vertical='center', wrap=False):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)

def _moeda(valor: float) -> str:
    return f'R$ {valor:,.2f}'

def _pct(valor: float) -> str:
    return f'{valor:.1f}%'


# ---------------------------------------------------------------------------
# Estilo de cabeçalho de tabela
# ---------------------------------------------------------------------------

def _aplicar_header(ws, linha: int, colunas: list[str]):
    for col_idx, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=linha, column=col_idx, value=titulo)
        cell.font      = _fonte(bold=True, cor=COR_BRANCO, tamanho=10)
        cell.fill      = _fill(COR_CINZA_HEADER)
        cell.alignment = _alinhar('center')
        cell.border    = _borda_fina()


def _aplicar_linha(ws, linha: int, valores: list, alternar: bool = False,
                   cor_fundo: str | None = None):
    fundo = cor_fundo or (COR_LINHA_ALT if alternar else COR_BRANCO)
    for col_idx, valor in enumerate(valores, start=1):
        cell = ws.cell(row=linha, column=col_idx, value=valor)
        cell.font      = _fonte(tamanho=9)
        cell.fill      = _fill(fundo)
        cell.alignment = _alinhar('left')
        cell.border    = _borda_fina()


def _largura_colunas(ws, larguras: list[int]):
    for idx, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# ---------------------------------------------------------------------------
# Aba 1 — Resumo
# ---------------------------------------------------------------------------

def _aba_resumo(wb: Workbook, resultado: ResultadoComparacao, gerado_em: str):
    ws = wb.active
    ws.title = 'Resumo'
    ws.sheet_view.showGridLines = False

    r = resultado.resumo

    # --- título ---
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value     = '📋 CONCILIAÇÃO DE NOTAS FISCAIS'
    c.font      = _fonte(bold=True, tamanho=16, cor=COR_BRANCO)
    c.fill      = _fill(COR_CINZA_HEADER)
    c.alignment = _alinhar('center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value     = f'Gerado em: {gerado_em}'
    c.font      = _fonte(tamanho=9, cor='777777')
    c.alignment = _alinhar('center')
    ws.row_dimensions[2].height = 18

    # --- cartões de KPI ---
    def _kpi(linha, col, titulo, valor, cor_fundo, cor_texto='FFFFFF'):
        col_l = get_column_letter(col)
        col_r = get_column_letter(col + 1)
        ws.merge_cells(f'{col_l}{linha}:{col_r}{linha}')
        ws.merge_cells(f'{col_l}{linha+1}:{col_r}{linha+1}')
        t = ws.cell(row=linha,   column=col, value=titulo)
        v = ws.cell(row=linha+1, column=col, value=valor)
        for cell in [t, v]:
            cell.fill      = _fill(cor_fundo)
            cell.alignment = _alinhar('center')
            cell.border    = _borda_fina()
        t.font = _fonte(bold=False, tamanho=9,  cor=cor_texto)
        v.font = _fonte(bold=True,  tamanho=16, cor=cor_texto)
        ws.row_dimensions[linha].height   = 18
        ws.row_dimensions[linha+1].height = 36

    linha_kpi = 4
    _kpi(linha_kpi, 1, 'TOTAL PREFEITURA (A)',  r['total_fonte_a'],       COR_AZUL)
    _kpi(linha_kpi, 3, 'TOTAL SISTEMA (B)',      r['total_fonte_b'],       COR_AZUL)
    _kpi(linha_kpi, 5, '✅ CONCILIADAS',         r['total_conciliadas'],   COR_VERDE_ESCURO)

    linha_kpi2 = linha_kpi + 3
    _kpi(linha_kpi2, 1, '⚠️  DIVERGENTES',       r['total_divergentes'],   COR_AMARELO, '2C3E50')
    _kpi(linha_kpi2, 3, '🔴 SÓ PREFEITURA',      r['total_so_prefeitura'], COR_VERMELHO)
    _kpi(linha_kpi2, 5, '🔵 SÓ SISTEMA',         r['total_so_sistema'],    COR_AZUL)

    # --- tabela de valores ---
    linha_tab = linha_kpi2 + 4
    ws.merge_cells(f'A{linha_tab}:F{linha_tab}')
    c = ws.cell(row=linha_tab, column=1, value='VALORES ENVOLVIDOS')
    c.font      = _fonte(bold=True, tamanho=11, cor=COR_BRANCO)
    c.fill      = _fill(COR_CINZA_HEADER)
    c.alignment = _alinhar('center')
    ws.row_dimensions[linha_tab].height = 22
    linha_tab += 1

    _aplicar_header(ws, linha_tab, ['Categoria', 'Qtd. NFs', 'Valor Total (R$)', 'Percentual', '', ''])
    linha_tab += 1

    linhas_valores = [
        ('Total Prefeitura (A)',  r['total_fonte_a'],        r['valor_total_a'],        '-'),
        ('Total Sistema (B)',     r['total_fonte_b'],        r['valor_total_b'],        '-'),
        ('✅ Conciliadas',        r['total_conciliadas'],    r['valor_conciliadas_a'],  _pct(r['pct_conciliadas'])),
        ('⚠️  Divergentes',       r['total_divergentes'],    r['valor_divergentes_a'],  _pct(r['pct_divergentes'])),
        ('🔴 Só Prefeitura',      r['total_so_prefeitura'],  r['valor_so_prefeitura'],  '-'),
        ('🔵 Só Sistema',         r['total_so_sistema'],     r['valor_so_sistema'],     '-'),
    ]

    cores_linhas = [
        COR_BRANCO, COR_LINHA_ALT,
        COR_VERDE_CLARO, COR_AMARELO_CLARO,
        COR_VERMELHO_CLARO, COR_AZUL_CLARO,
    ]

    for i, (cat, qtd, val, pct_str) in enumerate(linhas_valores):
        cor = cores_linhas[i]
        _aplicar_linha(ws, linha_tab, [cat, qtd, _moeda(val), pct_str, '', ''], cor_fundo=cor)
        ws.cell(row=linha_tab, column=1).font = _fonte(bold=True, tamanho=9)
        linha_tab += 1

    # --- nota de divergências ---
    linha_tab += 1
    ws.merge_cells(f'A{linha_tab}:F{linha_tab}')
    c = ws.cell(row=linha_tab, column=1,
                value=f'Divergências críticas: {r["total_divergencias_criticas"]}   |   '
                      f'Divergências leves: {r["total_divergencias_leves"]}')
    c.font      = _fonte(tamanho=9, cor='555555')
    c.alignment = _alinhar('center')

    _largura_colunas(ws, [28, 16, 22, 14, 14, 14])


# ---------------------------------------------------------------------------
# Aba 2 — Conciliadas
# ---------------------------------------------------------------------------

def _aba_conciliadas(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('Conciliadas')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value     = '✅ NFs CONCILIADAS — presentes em ambas as fontes, sem divergência'
    c.font      = _fonte(bold=True, tamanho=11, cor=COR_BRANCO)
    c.fill      = _fill(COR_VERDE_ESCURO)
    c.alignment = _alinhar('center')
    ws.row_dimensions[1].height = 24

    headers = ['NF', 'Valor (A=B)', 'Data (A=B)', 'CPF/CNPJ', 'Tomador (A)', 'Tomador (B)', 'Status']
    _aplicar_header(ws, 2, headers)

    if df.empty:
        ws.merge_cells('A3:G3')
        ws['A3'].value = 'Nenhuma NF conciliada encontrada.'
    else:
        for i, (_, row) in enumerate(df.iterrows()):
            vals = [
                row['nf'],
                _moeda(row['valor_a']),
                str(row['data_a']),
                row['cpf_cnpj_a'],
                row['tomador_a'],
                row['tomador_b'],
                '✅ OK',
            ]
            _aplicar_linha(ws, i + 3, vals, alternar=(i % 2 == 1), cor_fundo=COR_VERDE_CLARO if i % 2 == 0 else COR_BRANCO)

    _largura_colunas(ws, [10, 16, 14, 18, 32, 32, 10])


# ---------------------------------------------------------------------------
# Aba 3 — Divergentes
# ---------------------------------------------------------------------------

def _aba_divergentes(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('Divergentes')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:L1')
    c = ws['A1']
    c.value     = '⚠️  NFs DIVERGENTES — presentes em ambas as fontes, com diferença em algum campo'
    c.font      = _fonte(bold=True, tamanho=11, cor=COR_BRANCO)
    c.fill      = _fill(COR_AMARELO)
    c.alignment = _alinhar('center')
    ws.row_dimensions[1].height = 24

    headers = [
        'NF', 'Classificação',
        'Valor A', 'Valor B', 'Δ Valor', 'St. Valor',
        'Data A', 'Data B', 'Δ Dias', 'St. Data',
        'CPF/CNPJ A', 'CPF/CNPJ B', 'St. Doc',
        'Tomador A',
    ]
    _aplicar_header(ws, 2, headers)

    if df.empty:
        ws.merge_cells('A3:N3')
        ws['A3'].value = 'Nenhuma NF divergente encontrada.'
    else:
        for i, (_, row) in enumerate(df.iterrows()):
            is_critica = row['classificacao'] == STATUS_DIVERGENCIA_CRITICA
            cor = COR_VERMELHO_CLARO if is_critica else COR_AMARELO_CLARO

            emoji = '🔴' if is_critica else '⚠️'
            vals = [
                row['nf'],
                f'{emoji} {row["classificacao"]}',
                _moeda(row['valor_a']),
                _moeda(row['valor_b']),
                _moeda(row['diff_valor']),
                row['status_valor'],
                str(row['data_a']),
                str(row['data_b']),
                row['diff_dias'] if row['diff_dias'] is not None else '-',
                row['status_data'],
                row['cpf_cnpj_a'],
                row['cpf_cnpj_b'],
                row['status_doc'],
                row['tomador_a'],
            ]
            _aplicar_linha(ws, i + 3, vals, cor_fundo=cor)

    _largura_colunas(ws, [10, 24, 16, 16, 12, 10, 14, 14, 8, 10, 18, 18, 10, 32])


# ---------------------------------------------------------------------------
# Aba 4 — Só Prefeitura
# ---------------------------------------------------------------------------

def _aba_so_prefeitura(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('Só Prefeitura')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value     = '🔴 SÓ NA PREFEITURA — NFs que existem na Prefeitura mas não estão no Sistema'
    c.font      = _fonte(bold=True, tamanho=11, cor=COR_BRANCO)
    c.fill      = _fill(COR_VERMELHO)
    c.alignment = _alinhar('center')
    ws.row_dimensions[1].height = 24

    headers = ['NF', 'Valor', 'Data', 'CPF/CNPJ', 'Tipo Doc', 'Tomador']
    _aplicar_header(ws, 2, headers)

    if df.empty:
        ws.merge_cells('A3:F3')
        ws['A3'].value = 'Nenhuma NF exclusiva da Prefeitura.'
    else:
        for i, (_, row) in enumerate(df.iterrows()):
            tipo = row.get('tipo_doc', '')
            vals = [
                row['nf'],
                _moeda(row['valor']),
                str(row['data']),
                row['cpf_cnpj'],
                tipo,
                row['tomador'],
            ]
            cor = COR_VERMELHO_CLARO if i % 2 == 0 else COR_BRANCO
            _aplicar_linha(ws, i + 3, vals, cor_fundo=cor)

    _largura_colunas(ws, [10, 16, 14, 18, 12, 36])


# ---------------------------------------------------------------------------
# Aba 5 — Só Sistema
# ---------------------------------------------------------------------------

def _aba_so_sistema(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet('Só Sistema')
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value     = '🔵 SÓ NO SISTEMA — NFs que estão no Sistema mas não aparecem na Prefeitura'
    c.font      = _fonte(bold=True, tamanho=11, cor=COR_BRANCO)
    c.fill      = _fill(COR_AZUL)
    c.alignment = _alinhar('center')
    ws.row_dimensions[1].height = 24

    headers = ['NF', 'Valor', 'Data', 'CPF/CNPJ', 'Tipo Doc', 'Tomador']
    _aplicar_header(ws, 2, headers)

    if df.empty:
        ws.merge_cells('A3:F3')
        ws['A3'].value = 'Nenhuma NF exclusiva do Sistema.'
    else:
        for i, (_, row) in enumerate(df.iterrows()):
            tipo = row.get('tipo_doc', '')
            vals = [
                row['nf'],
                _moeda(row['valor']),
                str(row['data']),
                row['cpf_cnpj'],
                tipo,
                row['tomador'],
            ]
            cor = COR_AZUL_CLARO if i % 2 == 0 else COR_BRANCO
            _aplicar_linha(ws, i + 3, vals, cor_fundo=cor)

    _largura_colunas(ws, [10, 16, 14, 18, 12, 36])


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def gerar_excel(
    resultado: ResultadoComparacao,
    caminho_saida: str | Path | None = None,
) -> Path:
    """
    Gera o relatório Excel da conciliação.

    Parâmetros
    ----------
    resultado      : ResultadoComparacao (saída de comparator.comparar())
    caminho_saida  : caminho do arquivo de saída (opcional).
                     Padrão: 'conciliacao_YYYYMMDD_HHMMSS.xlsx' na pasta atual.

    Retorna
    -------
    Path do arquivo gerado.
    """
    gerado_em = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if caminho_saida is None:
        caminho_saida = Path(f'conciliacao_{timestamp}.xlsx')
    caminho_saida = Path(caminho_saida)

    wb = Workbook()

    _aba_resumo(wb, resultado, gerado_em)
    _aba_conciliadas(wb, resultado.df_conciliadas)
    _aba_divergentes(wb, resultado.df_divergentes)
    _aba_so_prefeitura(wb, resultado.df_so_prefeitura)
    _aba_so_sistema(wb, resultado.df_so_sistema)

    wb.save(str(caminho_saida))
    return caminho_saida


# ---------------------------------------------------------------------------
# CLI para teste rápido
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    import sys
    from reader import ler_arquivo
    from normalizer import normalizar
    from comparator import comparar

    if len(sys.argv) < 3:
        print('Uso: python report_excel.py <arquivo_A> <arquivo_B> [saida.xlsx]')
        sys.exit(1)

    df_a, _ = ler_arquivo(sys.argv[1], 'A')
    df_b, _ = ler_arquivo(sys.argv[2], 'B')

    res_a = normalizar(df_a, 'A')
    res_b = normalizar(df_b, 'B')

    resultado = comparar(res_a.df, res_b.df)

    saida = sys.argv[3] if len(sys.argv) > 3 else None
    arquivo = gerar_excel(resultado, saida)
    print(f'\n✅ Relatório gerado: {arquivo}')
